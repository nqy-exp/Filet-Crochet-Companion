import cv2
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import tempfile
import traceback

class PatternDigitizer:
    def __init__(self, image_path):
        self.image_path = image_path
        self.img = cv2.imread(image_path)
        if self.img is None:
            raise ValueError(f"Could not read image: {image_path}")
        self.rows = 0
        self.cols = 0
        self.matrix = [] # Store state: 0=White, 1=Black

    def process(self):
        """Core processing workflow"""
        print(f"Processing image: {self.image_path}")
        # 1. Image Pre-processing (Grayscale -> Binary)
        gray = cv2.cvtColor(self.img, cv2.COLOR_BGR2GRAY)
        binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                      cv2.THRESH_BINARY_INV, 11, 5)

        # 2. Grid Line Detection
        print("Detecting grid lines...")
        h_lines = self._get_lines(binary, axis=0)
        v_lines = self._get_lines(binary, axis=1)

        if len(h_lines) < 2 or len(v_lines) < 2:
            raise ValueError("Incomplete grid lines detected. Please check image quality.")

        self.rows = len(h_lines) - 1
        self.cols = len(v_lines) - 1
        print(f"Detection successful! Rows: {self.rows}, Cols: {self.cols}")

        # 3. Pixel Sampling
        self.matrix = []
        THRESHOLD = 0.3

        for r in range(self.rows):
            row_data = []
            for c in range(self.cols):
                y1, y2 = h_lines[r], h_lines[r+1]
                x1, x2 = v_lines[c], v_lines[c+1]
                cell_roi = binary[y1:y2, x1:x2]

                if cell_roi.size == 0:
                    row_data.append(0)
                    continue

                pixel_density = np.sum(cell_roi == 255) / cell_roi.size
                if pixel_density > THRESHOLD:
                    row_data.append(1)
                else:
                    row_data.append(0)
            self.matrix.append(row_data)
        print("Scanning complete!")

    def _get_lines(self, binary, axis):
        """This method must be at the same indentation level as process()"""
        if axis == 0: # Horizontal lines
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (binary.shape[1] // 40, 1))
        else: # Vertical lines
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, binary.shape[0] // 40))
        
        detected = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel)
        contours, _ = cv2.findContours(detected, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        coords = []
        for cnt in contours:
            x, y, w, h = cv2.boundingRect(cnt)
            coords.append(y + h // 2 if axis == 0 else x + w // 2)
        return sorted(coords)

    def export_to_excel(self, output_path):
        """Convert matrix to Excel file"""
        print(f"Generating Excel: {output_path}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pattern"

        colors = {0: "FFFFFF", 1: "333333"}
        thin_border = Border(left=Side(style='thin', color="DDDDDD"), 
                            right=Side(style='thin', color="DDDDDD"), 
                            top=Side(style='thin', color="DDDDDD"), 
                            bottom=Side(style='thin', color="DDDDDD"))

        for r in range(self.rows):
            for c in range(self.cols):
                cell = ws.cell(row=r+1, column=c+1)
                state = self.matrix[r][c]
                cell.fill = PatternFill(start_color=colors[state], end_color=colors[state], fill_type="solid")
                cell.border = thin_border

        for c in range(1, self.cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 3
        for r in range(1, self.rows + 1):
            ws.row_dimensions[r].height = 15

        wb.save(output_path)
        print("🎉 Success! Excel generated.")


# --- Main Entry Point ---
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python digitizer.py <image_path>")
        sys.exit(1)
    else:
        input_img = sys.argv[1]
        base_name = os.path.splitext(os.path.basename(input_img))[0]
        output_xlsx = os.path.join(tempfile.gettempdir(), base_name + "_pattern.xlsx")

        try:
            digitizer = PatternDigitizer(input_img)
            digitizer.process()
            digitizer.export_to_excel(output_xlsx)
            sys.exit(0) # 成功退出
        except Exception as e:
            import traceback
            # 使用 stderr 输出错误，确保 Electron 的 errorMsg 能捕获到
            print(f"❌ Error: {e}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            sys.exit(1) # 失败退出
