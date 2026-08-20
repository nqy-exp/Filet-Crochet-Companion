import cv2
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys

class SmartPatternDigitizer:
    def __init__(self, image_path):
        self.image_path = image_path
        self.img = cv2.imread(image_path)
        if self.img is None:
            # 统一使用英文报错
            raise ValueError(f"Could not read image: {image_path}")
        self.rows = 0
        self.cols = 0
        self.matrix = []  # Store results: 0=White, 1=Pattern/Color
        self.roi_coords = None # Record detected ROI (x_min, y_min, x_max, y_max)


    def process(self, target_rows, target_cols, brightness_threshold): 
        current_threshold = brightness_threshold 

        # 1. Pre-processing
        gray = cv2.cvtColor(self.img, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (7, 7), 0)

        # 2. Automatic ROI Detection (Based on brightness changes)
        binary = cv2.adaptiveThreshold(
            blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
            cv2.THRESH_BINARY_INV, 11, 2
        )
        coords = cv2.findNonZero(binary)
        if coords is None:
            raise ValueError("No pattern detected.")
        x, y, w, h = cv2.boundingRect(coords)
        self.roi_coords = (x, y, x + w, y + h)

        # Log detection to stdout (Standard English logs)
        print(f"ROI Detected: X:{x} Y:{y} W:{w} H:{h}")

        # 3. Sampling Logic
        self.rows = target_rows
        self.cols = target_cols
        self.matrix = []

        x_min, y_min, x_max, y_max = self.roi_coords
        cell_w_pixel = w / target_cols
        cell_h_pixel = h / target_rows

        for r in range(target_rows):
            row_data = []
            for c in range(target_cols):
                start_x = int(x_min + c * cell_w_pixel)
                end_x = int(x_min + (c + 1) * cell_w_pixel)
                start_y = int(y_min + r * cell_h_pixel)
                end_y = int(y_min + (r + 1) * cell_h_pixel)

                # Boundary check
                end_x = min(end_x, x_max)
                end_y = min(end_y, y_max)

                cell_roi = blurred[start_y:end_y, start_x:end_x]

                if cell_roi.size == 0:
                    row_data.append(0)
                    continue

                # Calculate average brightness
                avg_brightness = np.mean(cell_roi)

                # Use threshold passed from Frontend
                if avg_brightness < current_threshold:
                    row_data.append(1) # Pattern detected
                else:
                    row_data.append(0) # Background
            self.matrix.append(row_data)


    def export_to_excel(self, output_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Smart Pattern"

        # Color definition
        color_map = {
            0: "FFFFFF", # White
            1: "333333"  # Dark Grey/Black
        }
        thin_border = Border(
            left=Side(style='thin', color="DDDDDD"), 
            right=Side(style='thin', color="DDDDDD"), 
            top=Side(style='thin', color="DDDDDD"), 
            bottom=Side(style='thin', color="DDDDDD")
        )

        for r in range(self.rows):
            for c in range(self.cols):
                cell = ws.cell(row=r+1, column=c+1)
                state = self.matrix[r][c]
                fill_color = color_map[state]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = thin_border

        # Set square-like cells
        for c in range(1, self.cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 3
        for r in range(1, self.rows + 1):
            ws.row_dimensions[r].height = 15

        wb.save(output_path)
        print("🎉 Completed!")


# --- Main Entry Point ---
if __name__ == "__main__":
    import sys

    # Argument check (Usage instruction in English)
    if len(sys.argv) < 5:
        print("Usage: python smart_digitizer.py <image_path> <rows> <cols> <threshold>")
    else:
        input_img = sys.argv[1]
        try:
            target_r = int(sys.argv[2])
            target_c = int(sys.argv[3])
            user_threshold = int(sys.argv[4]) 
        except ValueError:
            sys.exit(1)

        output_xlsx = os.path.splitext(input_img)[0] + "_smart_pattern.xlsx"

        try:
            digitizer = SmartPatternDigitizer(input_img)
            digitizer.process(target_r, target_c, user_threshold) 
            digitizer.export_to_excel(output_xlsx)
        except Exception as e:
            import traceback
            # Output error in English so Electron can capture it
            print(f"❌ Process Error: {e}")
            traceback.print_exc()
